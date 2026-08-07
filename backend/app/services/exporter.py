from typing import List
import openpyxl
import os
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR

def save_to_excel(businesses: List[dict], business_type: str, location:str) -> dict:
    if OUTPUT_DIR is None:
        raise ValueError("OUTPUT_DIR is not configured")
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Clean up filename strings to avoid invalid characters
    clean_type = business_type.strip().replace(" ", "_")
    filename = f"leads_{clean_type}"

    headers = ["Business Name", "Website", "Phone", "Address", "Email"]

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Leads Data")
        else:
            ws.title = "Leads Data"

        ws.append(headers)

        # Map dictionary keys explicitly to match the header column order
        for item in businesses:
            row_data = [
                item.get("name", ""),
                item.get("website", ""),
                item.get("phone", ""),
                item.get("address", ""),
                item.get("email", "")
            ]
            ws.append(row_data)

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_len = 0
            col_idx = col[0].column
            if col_idx is None:
                continue
            col_letter = get_column_letter(col_idx)
            
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
                    
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        file_path = os.path.join(OUTPUT_DIR, f"{filename}.xlsx")
        wb.save(file_path)
        print(f"Excel file created successfully at: {file_path}")
        return {
            "file_name": file_path,
            "total_businesses": len(businesses),
            "business_type": business_type,
            "location": location,
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {}